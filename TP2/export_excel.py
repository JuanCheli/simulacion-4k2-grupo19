import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from interfaz import *
from utils import *
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice


def exportar_a_excel(app):

    # Verificar que hay datos para exportar
    if app.datos_generados is None:
        messagebox.showerror("Error", "No hay datos para exportar. Genera una muestra primero.")
        return

    # Pedir al usuario dónde guardar el archivo
    archivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Guardar como Excel"
    )

    if not archivo:  # Si el usuario cancela, salir
        return

    # Crear un libro de trabajo y hojas
    wb = openpyxl.Workbook()

    # Hoja 1: Datos brutos
    ws_datos = wb.active
    ws_datos.title = "Datos Generados"

    # Hoja 2: Histograma (gráfico y datos)
    ws_hist = wb.create_sheet("Histograma")

    # Hoja 3: Tabla de frecuencias
    ws_freq = wb.create_sheet("Tabla de Frecuencias")

    # --- Hoja 1: Datos Generados ---
    ws_datos = _escribir_datos_generados(ws_datos, app)

    # --- Hoja 2: Histograma (datos y gráfico) ---
    ws_hist = _escribir_datos_histograma(ws_hist, app)

    # --- Hoja 3: Tabla de Frecuencias ---
    ws_freq = _escribir_tabla_frecuencias(ws_freq, app)

    # Ajustar anchos de columna automáticamente en todas las hojas
    for ws in [ws_datos, ws_hist, ws_freq]:
        _ajustar_ancho_columnas(ws)

    # Guardar el archivo
    wb.save(archivo)

    messagebox.showinfo("Exportación exitosa",
                        f"Los datos han sido exportados con éxito a:\n{os.path.basename(archivo)}")


def _escribir_datos_generados(ws, app):
    ws.cell(1, 1, "Datos Generados")
    ws.cell(1, 1).font = Font(bold=True, size=14)

    # Información de la distribución
    distribucion = app.distribucion_var.get()

    ws.cell(3, 1, "Distribución:")
    ws.cell(3, 2, distribucion.capitalize())

    fila = 4
    if distribucion == "uniforme":
        a = float(app.uniforme_a_var.get())
        b = float(app.uniforme_b_var.get())
        ws.cell(fila, 1, "Valor mínimo (a):")
        ws.cell(fila, 2, a)
        fila += 1
        ws.cell(fila, 1, "Valor máximo (b):")
        ws.cell(fila, 2, b)
    elif distribucion == "exponencial":
        lambda_val = float(app.exponencial_lambda_var.get())
        ws.cell(fila, 1, "Lambda (λ):")
        ws.cell(fila, 2, lambda_val)
    elif distribucion == "normal":
        media = float(app.normal_media_var.get())
        desviacion = float(app.normal_desviacion_var.get())
        ws.cell(fila, 1, "Media (μ):")
        ws.cell(fila, 2, media)
        fila += 1
        ws.cell(fila, 1, "Desviación estándar (σ):")
        ws.cell(fila, 2, desviacion)

    fila += 2
    ws.cell(fila, 1, "Tamaño de muestra:")
    ws.cell(fila, 2, len(app.datos_generados))

    # Estadísticas básicas
    fila += 2
    ws.cell(fila, 1, "Estadísticas Básicas")
    ws.cell(fila, 1).font = Font(bold=True)
    fila += 1

    media = np.mean(app.datos_generados)
    mediana = np.median(app.datos_generados)
    desv_est = np.std(app.datos_generados)
    min_valor = np.min(app.datos_generados)
    max_valor = np.max(app.datos_generados)

    ws.cell(fila, 1, "Media:")
    ws.cell(fila, 2, f"{media:.4f}")
    fila += 1
    ws.cell(fila, 1, "Mediana:")
    ws.cell(fila, 2, f"{mediana:.4f}")
    fila += 1
    ws.cell(fila, 1, "Desviación estándar:")
    ws.cell(fila, 2, f"{desv_est:.4f}")
    fila += 1
    ws.cell(fila, 1, "Valor mínimo:")
    ws.cell(fila, 2, f"{min_valor:.4f}")
    fila += 1
    ws.cell(fila, 1, "Valor máximo:")
    ws.cell(fila, 2, f"{max_valor:.4f}")

    # Datos brutos
    fila += 2
    ws.cell(fila, 1, "Serie de Números Generados")
    ws.cell(fila, 1).font = Font(bold=True)
    fila += 1

    # Encabezados
    columnas_por_fila = 10
    for j in range(columnas_por_fila):
        ws.cell(fila, j + 1, f"Valor {j + 1}")
        ws.cell(fila, j + 1).font = Font(bold=True)

    # Datos
    fila += 1
    for i in range(0, len(app.datos_generados), columnas_por_fila):
        col = 1
        for j in range(columnas_por_fila):
            if i + j < len(app.datos_generados):
                ws.cell(fila, col, f"{app.datos_generados[i + j]:.4f}")
            col += 1
        fila += 1
    return ws

def _escribir_datos_histograma(ws, app):
    from openpyxl.chart import BarChart, Reference

    ws.cell(1, 1, "Datos del Histograma")
    ws.cell(1, 1).font = Font(bold=True, size=14)

    num_intervalos = int(app.num_intervalos_var.get())

    # Crear histograma
    frecuencias, bordes = np.histogram(app.datos_generados, bins=num_intervalos)

    ws.cell(3, 1, "Número de intervalos:")
    ws.cell(3, 2, num_intervalos)

    # Encabezados para los datos del histograma
    fila = 5
    ws.cell(fila, 1, "Intervalo")
    ws.cell(fila, 2, "Límite inferior")
    ws.cell(fila, 3, "Límite superior")
    ws.cell(fila, 4, "Marca de clase")
    ws.cell(fila, 5, "Frecuencia")
    for col in range(1, 6):
        ws.cell(fila, col).font = Font(bold=True)

    # Datos del histograma
    for i in range(len(frecuencias)):
        fila += 1
        ws.cell(fila, 1, i + 1)
        ws.cell(fila, 2, f"{bordes[i]:.4f}")
        ws.cell(fila, 3, f"{bordes[i + 1]:.4f}")
        marca = (bordes[i] + bordes[i + 1]) / 2
        ws.cell(fila, 4, f"{marca:.4f}")
        ws.cell(fila, 5, frecuencias[i])

    # Crear un gráfico de barras básico
    try:
        chart = BarChart()
        chart.title = "Histograma"
        chart.x_axis.title = "Marca de clase"
        chart.y_axis.title = "Frecuencia"

        # Referencia a las frecuencias (los valores de las barras)
        data = Reference(ws, min_col=5, min_row=5, max_row=fila, max_col=5)

        # Referencia a las marcas de clase (las etiquetas del eje x)
        categories = Reference(ws, min_col=4, min_row=6, max_row=fila, max_col=4)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Agregar el gráfico a la hoja
        ws.add_chart(chart, "H2")
    except Exception as e:
        # Si hay un error al generar el gráfico, añadir una nota
        ws.cell(3, 4, "Error al generar gráfico:")
        ws.cell(3, 5, str(e))

    # Estadísticas básicas
    fila_stats = fila + 3
    ws.cell(fila_stats, 1, "Estadísticas Básicas")
    ws.cell(fila_stats, 1).font = Font(bold=True)
    fila_stats += 1

    # Calcular estadísticas
    media = np.mean(app.datos_generados)
    mediana = np.median(app.datos_generados)
    desv_est = np.std(app.datos_generados)
    min_valor = np.min(app.datos_generados)
    max_valor = np.max(app.datos_generados)

    # Mostrar estadísticas en una fila
    estadisticas = [
        f"Media: {media:.4f}",
        f"Mediana: {mediana:.4f}",
        f"Desv. Estándar: {desv_est:.4f}",
        f"Mín: {min_valor:.4f}",
        f"Máx: {max_valor:.4f}"
    ]

    for i, stat in enumerate(estadisticas):
        ws.cell(fila_stats, i + 1, stat)

    return ws


def _escribir_tabla_frecuencias(ws, app):
    ws.cell(1, 1, "Tabla de Frecuencias")
    ws.cell(1, 1).font = Font(bold=True, size=14)

    tabla_datos = _calcular_tabla_frecuencias(app)

    # Encabezados
    fila = 3
    ws.cell(fila, 1, f"Tabla de Frecuencias ({len(tabla_datos['frecuencias'])} intervalos tras agrupamiento)")
    ws.cell(fila, 1).font = Font(bold=True)

    fila += 2
    cols = ["#", "Limite inferior", "Limite superior", "Marca de clase",
            "Frecuencia observada", "Frecuencia esperada"]
    for c, h in enumerate(cols):
        ws.cell(fila, c + 1, h)
        ws.cell(fila, c + 1).font = Font(bold=True)
        ws.cell(fila, c + 1).alignment = Alignment(horizontal='center')
        ws.cell(fila, c + 1).fill = PatternFill("solid", fgColor="DDDDDD")

    # Datos de la tabla
    for i in range(len(tabla_datos['frecuencias'])):
        fila += 1
        ws.cell(fila, 1, i + 1)
        ws.cell(fila, 2, f"{tabla_datos['bordes'][i]:.4f}")
        ws.cell(fila, 3, f"{tabla_datos['bordes'][i + 1]:.4f}")
        marca = (tabla_datos['bordes'][i] + tabla_datos['bordes'][i + 1]) / 2
        ws.cell(fila, 4, f"{marca:.4f}")
        ws.cell(fila, 5, tabla_datos['frecuencias'][i])
        ws.cell(fila, 6, f"{tabla_datos['f_esperadas'][i]:.2f}")

    # Resultado prueba Chi-cuadrado
    fila += 2
    texto = f"χ²={tabla_datos['chi2']:.3f}  df={tabla_datos['df']}  χ²₀.₀₅={tabla_datos['crit']:.3f} → {tabla_datos['decision']}"
    ws.cell(fila, 1, texto)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    ws.cell(fila, 1).font = Font(bold=True)
    ws.cell(fila, 1).alignment = Alignment(horizontal='center')

    return ws


def _calcular_tabla_frecuencias(app):
    datos = app.datos_generados
    num_intervalos = int(app.num_intervalos_var.get())
    distribucion = app.distribucion_var.get()

    # Histograma y listas mutables
    frecuencias, bordes = np.histogram(datos, bins=num_intervalos)
    frecuencias = frecuencias.tolist()
    bordes = bordes.tolist()

    n = len(datos)

    # Frecuencias esperadas iniciales
    f_esperadas = []
    for i in range(len(frecuencias)):
        a, b = bordes[i], bordes[i + 1]
        if distribucion == "uniforme":
            a0, b0 = float(app.uniforme_a_var.get()), float(app.uniforme_b_var.get())
            p = (b - a) / (b0 - a0)
        elif distribucion == "exponencial":
            lam = float(app.exponencial_lambda_var.get())
            p = cdf_exponencial(b, lam) - cdf_exponencial(a, lam)
        else:  # normal
            mu, sigma = float(app.normal_media_var.get()), float(app.normal_desviacion_var.get())
            p = cdf_normal(b, mu, sigma) - cdf_normal(a, mu, sigma)
        f_esperadas.append(p * n)

    # Función auxiliar de agrupamiento
    def agrupar(i, j):
        i0, j0 = sorted((i, j))
        bordes[i0 + 1] = bordes[j0 + 1]
        frecuencias[i0] += frecuencias[j0]
        f_esperadas[i0] += f_esperadas[j0]
        del bordes[j0 + 1], frecuencias[j0], f_esperadas[j0]

    # Agrupar intervalos con fe < 5
    changed = True
    while changed:
        changed = False
        for idx, fe in enumerate(f_esperadas):
            if fe < 5:
                if idx == 0:
                    vecino = 1
                elif idx == len(f_esperadas) - 1:
                    vecino = idx - 1
                else:
                    vecino = idx + 1 if f_esperadas[idx + 1] < f_esperadas[idx - 1] else idx - 1
                agrupar(idx, vecino)
                changed = True
                break

    # Calcular final de χ² y decisión
    chi2 = sum((o - e) ** 2 / e for o, e in zip(frecuencias, f_esperadas))
    if distribucion == "uniforme":
        df = len(frecuencias) - 1
    elif distribucion == "exponencial":
        df = len(frecuencias) - 2
    else:
        df = len(frecuencias) - 3
    crit = CHI2_CRITICOS_005.get(df, float('nan'))
    decision = "Rechazar H₀" if chi2 > crit else "No rechazar H₀"

    return {
        "bordes": bordes,
        "frecuencias": frecuencias,
        "f_esperadas": f_esperadas,
        "chi2": chi2,
        "df": df,
        "crit": crit,
        "decision": decision
    }


def _ajustar_ancho_columnas(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width